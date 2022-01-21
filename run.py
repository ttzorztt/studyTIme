import time
import openpyxl
def creatExcel(savePath:str):
    myExcel = openpyxl.Workbook()
    myExcel.create_sheet('All')
    sheet = myExcel['Sheet']
    myExcel.remove(sheet)
    sheet = myExcel['All']
    sheetTitle = ["","data","math","english","politics","computer","allTime"]
    for i in range(1,7):
        sheet.cell(row=1,column=i).value = sheetTitle[i]
    myExcel.save(savePath)

def InputObject(insertString:str) -> int:
    while(1):
        print("please input you want {} object,\n math,english,computer or politics?".format(insertString))
        study_object = input()
        if study_object != 'math' and study_object != 'english' and study_object != 'computer' and study_object != 'politics':
            print("object is wrong,please repeat the operation")
        else:
            if study_object == 'math':
                return 2
            elif study_object == 'english':
                return 3
            elif study_object == 'politics':
                return 4
            elif study_object == 'computer':
                return 5

def showStudyTime():
    print()
    myWorkBook = openpyxl.load_workbook("/home/tmd/Documents/note/study_database.xlsx")
    mysheet = myWorkBook['All']
    for i in range(1,7):
        print("{}\t".format(mysheet.cell(row=mysheet.max_row, column = i).value),end='')
    myWorkBook.save("/home/tmd/Documents/note/study_database.xlsx")
    print()
def start_study():
    myWorkBook = openpyxl.load_workbook("/home/tmd/Documents/note/study_database.xlsx")
    mysheet = myWorkBook['All']
    if mysheet.cell(row = 1, column = 7).value != None:
        print("please stop the before study")
        return
    object = InputObject("study")
    mysheet.cell(row = 1, column = 8).value = round(time.time() / 60)
    mysheet.cell(row = 1, column = 7).value = object
    myWorkBook.save("/home/tmd/Documents/note/study_database.xlsx")

def addStudyTime():
    object = InputObject("add")
    addStudyTime = input("add the studyTime is: ")
    myWorkBook = openpyxl.load_workbook("/home/tmd/Documents/note/study_database.xlsx")
    mysheet = myWorkBook['All']
    mysheet.cell(row = mysheet.max_row, column = object).value = mysheet.cell(row = mysheet.max_row, column = object).value + int(addStudyTime)
    mysheet.cell(row = mysheet.max_row, column = 6).value = mysheet.cell(row = mysheet.max_row, column = 6).value + int(addStudyTime)
    print("this stduy time is {}".format(addStudyTime))
    print("this day study time is {}".format(mysheet.cell(row = mysheet.max_row,column = 6).value))
    myWorkBook.save("/home/tmd/Documents/note/study_database.xlsx")

def end_stduy():
    myWorkBook = openpyxl.load_workbook("/home/tmd/Documents/note/study_database.xlsx")
    mysheet = myWorkBook['All']
    if mysheet.cell(row = 1, column = 7).value == None:
        print("please start study before")
        return
    object = mysheet.cell(row = 1, column = 7).value
    studyTime = round(time.time() / 60) - mysheet.cell(row = 1,column = 8).value
    mysheet.cell(row = mysheet.max_row,column = object).value = mysheet.cell(row = mysheet.max_row,column = object).value + studyTime
    mysheet.cell(row = mysheet.max_row, column = 6).value = mysheet.cell(row = mysheet.max_row, column = 6).value + studyTime
    mysheet.cell(row = 1, column = 7).value = None
    mysheet.cell(row = 1, column = 8).value = None
    myWorkBook.save("/home/tmd/Documents/note/study_database.xlsx")
    print("this stduy time is {}".format(studyTime))
    print("this day study time is {}".format(mysheet.cell(row = mysheet.max_row,column = 6).value))
def isFile() -> bool:
    myExcelPath = "/home/tmd/Documents/note/study_database.xlsx"
    try:
        file = open(myExcelPath)
    except FileNotFoundError:
        while(1):
            print("The file that named study_database.xlsx is not found!")
            test = input("Create or not? (y/n)")
            if test == 'y' or test == 'Y':
                creatExcel(myExcelPath)
                return True
            elif test == 'n' or test == 'N':
                print("OK,the project is over")
                return False
            else:
                print("input is wrong, please repeat the operation")
                
if __name__ == "__main__":
    print(time.ctime())
    if isFile() == False:
        exit(0)
    myWorkBook = openpyxl.load_workbook("/home/tmd/Documents/note/study_database.xlsx")
    mysheet = myWorkBook['All']
    today = time.ctime()[4:10]
    if mysheet.cell(row = mysheet.max_row, column = 1).value[:6] != today:
        mysheet.cell(row = mysheet.max_row + 1, column = 1).value = time.ctime()[4:]
        for i in range(2,7):
            mysheet.cell(row = mysheet.max_row, column = i).value = 0
    myWorkBook.save("/home/tmd/Documents/note/study_database.xlsx")
    while(1):
        startOrOver = input("Start,Over,add the time or show now study time? (Y/N/A/S)")
        if startOrOver == 'S' or startOrOver == 's':
            showStudyTime()
            break
        if startOrOver == 'A' or startOrOver == 'a':
            addStudyTime()
            break
        if startOrOver == 'Y' or startOrOver == 'y':
            start_study()
            break
        elif startOrOver == 'n' or startOrOver == 'N':
            end_stduy()
            break
        else:
            print("please input Y or N")
